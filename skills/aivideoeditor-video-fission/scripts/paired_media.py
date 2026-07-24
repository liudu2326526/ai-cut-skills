from __future__ import annotations

import argparse
import csv
import json
import os
import re
import shutil
import subprocess
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any


VIDEO_SUFFIXES = {".mp4", ".mov", ".mkv", ".avi"}
MEDIA_SUFFIXES = VIDEO_SUFFIXES | {".m4a", ".mp3", ".wav", ".aac", ".flac", ".ogg"}
TRAILING_SEQUENCE_RE = re.compile(r"\(\d+\)$")
MANIFEST_COLUMNS = [
    "mode",
    "output_name",
    "source_video",
    "audio_source",
    "variant_id",
    "deleted_frames",
    "cover_timestamp",
    "cover_quality_status",
    "cover_similarity_status",
    "combo_signature",
    "source_chain",
    "variation_status",
    "duplicate_risk",
    "quality_status",
    "business_tag",
    "material_type",
    "authorization_note",
    "upload_note",
    "creative_unit_id",
    "note",
]


@dataclass
class VariantPlan:
    variant_id: str
    source_video: Path
    audio_source: Path | None
    output_name: str
    output_path: Path
    combo_signature: str
    mode: str = "paired_media_match"
    quality_status: str = "ready"
    variation_status: str = "paired"
    duplicate_risk: str = "n/a"
    note: str = ""

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["source_video"] = str(self.source_video)
        data["audio_source"] = str(self.audio_source) if self.audio_source else ""
        data["output_path"] = str(self.output_path)
        return data


def main() -> int:
    args = parse_args()
    ffmpeg = resolve_binary("ffmpeg", args.ffmpeg)
    ffprobe = resolve_binary("ffprobe", args.ffprobe)
    input_folder = Path(args.input_folder).resolve()
    if not input_folder.is_dir():
        raise SystemExit(f"Input folder does not exist: {input_folder}")

    task_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    task_root = Path(args.output_root).resolve() / task_id
    videos_dir = task_root / "videos"
    videos_dir.mkdir(parents=True, exist_ok=True)

    scan = scan_pairable_media_folder(input_folder, ffprobe)
    variants: list[VariantPlan] = []
    failures: list[dict[str, str]] = []
    used_names: set[str] = set()
    for index, pair in enumerate(scan["pairs"], start=1):
        output_name = build_pair_output_name(pair["pair_id"], used_names)
        output_path = videos_dir / f"{output_name}.mp4"
        try:
            mode = render_paired_media_video(pair["video_path"], pair["audio_path"], output_path, ffmpeg)
            note = f"{pair.get('note', '')} | process_mode={mode}".strip(" |")
            variant = VariantPlan(
                variant_id=f"PAIR-{index:03d}",
                source_video=pair["video_path"],
                audio_source=pair["audio_path"],
                output_name=output_name,
                output_path=output_path,
                combo_signature=pair["pair_id"],
                note=note,
            )
        except Exception as exc:
            failures.append({
                "pair_id": pair["pair_id"],
                "video_source": str(pair["video_path"]),
                "audio_source": str(pair["audio_path"]),
                "error": str(exc),
            })
            variant = VariantPlan(
                variant_id=f"PAIR-{index:03d}",
                source_video=pair["video_path"],
                audio_source=pair["audio_path"],
                output_name=output_name,
                output_path=output_path,
                combo_signature=pair["pair_id"],
                quality_status="failed",
                variation_status="pair_failed",
                note=str(exc),
            )
        variants.append(variant)
        print(f"[{index}/{len(scan['pairs'])}] {variant.quality_status} {output_path}")

    payload = {
        "task_id": task_id,
        "mode": "paired_media_match",
        "config": vars(args),
        "scan_summary": scan["summary"],
        "scan_items": normalize_scan_items(scan["items"]),
        "unmatched": normalize_records(scan["unmatched"]),
        "conflicts": normalize_records(scan["conflicts"]),
        "failures": failures,
        "variants": [item.to_dict() for item in variants],
    }
    write_outputs(task_root, payload, variants)
    print(str(task_root))
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Pair pure video files with pure audio files.")
    parser.add_argument("--input-folder", required=True)
    parser.add_argument("--output-root", required=True)
    parser.add_argument("--task-name", default="pair_batch")
    parser.add_argument("--max-name-length", type=int, default=240)
    parser.add_argument("--ffmpeg", default="")
    parser.add_argument("--ffprobe", default="")
    return parser.parse_args()


def resolve_binary(name: str, explicit: str = "") -> str:
    candidates = [
        explicit,
        os.environ.get(f"{name.upper()}_BIN", ""),
        shutil.which(name) or "",
        shutil.which(f"{name}.exe") or "",
        str(Path(__file__).resolve().parent / "bin" / f"{name}.exe"),
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(Path(candidate))
    raise SystemExit(f"Missing dependency: {name}. Pass --{name} or set {name.upper()}_BIN.")


def run_process(args: list[str]) -> subprocess.CompletedProcess[str]:
    startupinfo = None
    creationflags = 0
    if hasattr(subprocess, "STARTUPINFO"):
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        creationflags = subprocess.CREATE_NO_WINDOW
    try:
        return subprocess.run(
            args,
            check=True,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            encoding="utf-8",
            errors="replace",
            startupinfo=startupinfo,
            creationflags=creationflags,
        )
    except subprocess.CalledProcessError as exc:
        raise RuntimeError((exc.stderr or exc.stdout or str(exc)).strip()) from exc


def scan_pairable_media_folder(folder: Path, ffprobe: str) -> dict[str, Any]:
    files = sorted(path for path in folder.iterdir() if path.is_file() and path.suffix.lower() in MEDIA_SUFFIXES)
    groups: dict[str, dict[str, list[Path]]] = {}
    items: list[dict[str, Any]] = []
    counts = {"video_only_count": 0, "audio_only_count": 0, "conflict_count": 0, "invalid_count": 0}
    for path in files:
        role = classify_media_file(path, ffprobe)
        pair_id = normalize_pair_id(path.stem)
        items.append({"path": path, "pair_id": pair_id, "role": role})
        groups.setdefault(pair_id, {"video_only": [], "audio_only": [], "av_both": [], "invalid": []})[role].append(path)
        if role == "video_only":
            counts["video_only_count"] += 1
        elif role == "audio_only":
            counts["audio_only_count"] += 1
        elif role == "av_both":
            counts["conflict_count"] += 1
        else:
            counts["invalid_count"] += 1

    pairs: list[dict[str, Any]] = []
    unmatched: list[dict[str, Any]] = []
    conflicts: list[dict[str, Any]] = []
    for pair_id in sorted(groups):
        grouped = groups[pair_id]
        videos = sorted(grouped["video_only"])
        audios = sorted(grouped["audio_only"])
        av_both = sorted(grouped["av_both"])
        invalids = sorted(grouped["invalid"])
        if av_both or invalids:
            conflicts.append({
                "pair_id": pair_id,
                "status": "conflict",
                "video_source": "",
                "audio_source": "",
                "output": "",
                "note": "; ".join([*(f"av_both:{path.name}" for path in av_both), *(f"invalid:{path.name}" for path in invalids)]),
            })
        if videos and audios:
            note_parts = []
            if len(videos) > 1:
                note_parts.append("extra videos ignored: " + ", ".join(path.name for path in videos[1:]))
            if len(audios) > 1:
                note_parts.append("extra audios ignored: " + ", ".join(path.name for path in audios[1:]))
            pairs.append({"pair_id": pair_id, "video_path": videos[0], "audio_path": audios[0], "note": "; ".join(note_parts)})
        elif videos or audios:
            unmatched.append({
                "pair_id": pair_id,
                "status": "unmatched",
                "video_source": videos[0].name if videos else "",
                "audio_source": audios[0].name if audios else "",
                "output": "",
                "note": "missing audio" if videos else "missing video",
            })
    return {
        "items": items,
        "pairs": pairs,
        "unmatched": unmatched,
        "conflicts": conflicts,
        "summary": {
            "total_files": len(files),
            "matched_pair_count": len(pairs),
            "unmatched_count": len(unmatched),
            **counts,
        },
    }


def classify_media_file(path: Path, ffprobe: str) -> str:
    result = run_process([ffprobe, "-v", "error", "-print_format", "json", "-show_streams", "-show_format", str(path)])
    data = json.loads(result.stdout)
    has_video = any(item.get("codec_type") == "video" for item in data.get("streams", []))
    has_audio = any(item.get("codec_type") == "audio" for item in data.get("streams", []))
    if has_video and not has_audio:
        return "video_only"
    if has_audio and not has_video:
        return "audio_only"
    if has_video and has_audio:
        return "av_both"
    return "invalid"


def normalize_pair_id(name: str) -> str:
    trimmed = TRAILING_SEQUENCE_RE.sub("", name.strip()).strip()
    return trimmed or name.strip() or "pair"


def build_pair_output_name(pair_id: str, used_names: set[str]) -> str:
    base = sanitize_windows_stem(pair_id, 240)
    candidate = base or "pair"
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate
    sequence = 1
    while True:
        suffix = f"-{sequence:02d}"
        candidate = f"{sanitize_windows_stem(base, 240 - len(suffix))}{suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        sequence += 1


def sanitize_windows_stem(name: str, max_length: int) -> str:
    invalid = set('<>:"/\\|?*')
    cleaned = "".join("_" if ch in invalid or ord(ch) < 32 else ch for ch in name.strip())
    cleaned = " ".join(cleaned.split())
    return cleaned[:max_length].rstrip(" ._-") or "variant"


def render_paired_media_video(video_path: Path, audio_path: Path, output_path: Path, ffmpeg: str) -> str:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        run_process([
            ffmpeg,
            "-y",
            "-i",
            str(video_path),
            "-i",
            str(audio_path),
            "-map",
            "0:v:0",
            "-map",
            "1:a:0",
            "-c:v",
            "copy",
            "-c:a",
            "copy",
            "-shortest",
            "-movflags",
            "+faststart",
            str(output_path),
        ])
        return "stream_copy"
    except RuntimeError:
        output_path.unlink(missing_ok=True)
        run_process([
            ffmpeg,
            "-y",
            "-i",
            str(video_path),
            "-i",
            str(audio_path),
            "-map",
            "0:v:0",
            "-map",
            "1:a:0",
            "-c:v",
            "libx264",
            "-preset",
            "fast",
            "-crf",
            "22",
            "-c:a",
            "aac",
            "-shortest",
            "-movflags",
            "+faststart",
            str(output_path),
        ])
        return "reencode_fallback"


def normalize_scan_items(items: list[dict[str, Any]]) -> list[dict[str, str]]:
    return [{"path": str(item["path"]), "pair_id": item["pair_id"], "role": item["role"]} for item in items]


def normalize_records(records: list[dict[str, Any]]) -> list[dict[str, str]]:
    normalized = []
    for record in records:
        normalized.append({key: str(value) for key, value in record.items()})
    return normalized


def write_outputs(task_root: Path, payload: dict[str, Any], variants: list[VariantPlan]) -> None:
    task_root.mkdir(parents=True, exist_ok=True)
    (task_root / "task.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_manifest_csv(task_root / "manifest.csv", variants)
    write_manifest_xlsx(task_root / "manifest.xlsx", variants)
    write_task_log(task_root / "task.log", payload)


def write_manifest_csv(path: Path, variants: list[VariantPlan]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MANIFEST_COLUMNS)
        writer.writeheader()
        writer.writerows(plan_to_row(item) for item in variants)


def write_manifest_xlsx(path: Path, variants: list[VariantPlan]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "manifest"
    sheet.append(MANIFEST_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for item in variants:
        row = plan_to_row(item)
        sheet.append([row[column] for column in MANIFEST_COLUMNS])
    workbook.save(path)


def plan_to_row(plan: VariantPlan) -> dict[str, str]:
    return {
        "mode": plan.mode,
        "output_name": plan.output_name,
        "source_video": str(plan.source_video),
        "audio_source": str(plan.audio_source or ""),
        "variant_id": plan.variant_id,
        "deleted_frames": "",
        "cover_timestamp": "",
        "cover_quality_status": "n/a",
        "cover_similarity_status": "n/a",
        "combo_signature": plan.combo_signature,
        "source_chain": "",
        "variation_status": plan.variation_status,
        "duplicate_risk": plan.duplicate_risk,
        "quality_status": plan.quality_status,
        "business_tag": "",
        "material_type": "",
        "authorization_note": "",
        "upload_note": "",
        "creative_unit_id": "",
        "note": plan.note,
    }


def write_task_log(path: Path, payload: dict[str, Any]) -> None:
    lines = [
        f"task_id: {payload.get('task_id', '')}",
        "mode: paired_media_match",
        f"created_at: {datetime.now().isoformat(timespec='seconds')}",
        f"scan_summary: {payload.get('scan_summary', {})}",
        "",
        "[variants]",
    ]
    for item in payload.get("variants", []):
        lines.append(f"{item.get('variant_id')} | pair={item.get('combo_signature')} | status={item.get('quality_status')} | {item.get('output_path')}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
