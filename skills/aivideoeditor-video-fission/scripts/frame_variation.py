from __future__ import annotations

import argparse
import csv
import hashlib
import json
import math
import os
import random
import shutil
import subprocess
import traceback
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any


VIDEO_SUFFIXES = {".mp4", ".mov", ".mkv", ".avi"}
MANIFEST_COLUMNS = [
    "mode",
    "output_name",
    "source_video",
    "audio_source",
    "variant_id",
    "deleted_frames",
    "cover_timestamp",
    "cover_quality_status",
    "cover_similarity_status",
    "combo_signature",
    "source_chain",
    "variation_status",
    "duplicate_risk",
    "quality_status",
    "business_tag",
    "material_type",
    "authorization_note",
    "upload_note",
    "creative_unit_id",
    "note",
]


@dataclass
class SourceVideo:
    path: Path
    duration: float
    width: int
    height: int
    fps: float
    frame_count: int
    has_audio: bool
    bit_rate: int = 0
    md5: str = ""


@dataclass
class CoverCandidate:
    timestamp: float
    path: Path
    sharpness: float = 0.0
    brightness: float = 0.0
    hash_value: str = ""


@dataclass
class VariantPlan:
    variant_id: str
    source_video: Path
    output_name: str
    output_path: Path
    mode: str = "frame_variation"
    deleted_frames: list[int] = field(default_factory=list)
    frame_signature: str = ""
    combo_signature: str = ""
    source_chain: list[Path] = field(default_factory=list)
    cover_timestamp: float = 0.0
    cover_quality_status: str = "pending"
    cover_similarity_status: str = "pending"
    quality_status: str = "ready"
    variation_status: str = "unique"
    duplicate_risk: str = "low"
    business_tag: str = ""
    material_type: str = ""
    authorization_note: str = ""
    upload_note: str = ""
    creative_unit_id: str = ""
    note: str = ""

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["source_video"] = str(self.source_video)
        data["output_path"] = str(self.output_path)
        data["source_chain"] = [str(path) for path in self.source_chain]
        return data


def main() -> int:
    args = parse_args()
    task_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_root = Path(args.output_root).resolve()
    task_root = output_root / task_id
    videos_dir = task_root / "videos"
    temp_cover_dir = task_root / "_cover_candidates"
    videos_dir.mkdir(parents=True, exist_ok=True)
    temp_cover_dir.mkdir(parents=True, exist_ok=True)
    run_log = task_root / "run.log"
    log_event(run_log, f"start frame_variation task_id={task_id}")
    log_event(run_log, f"config={json.dumps(vars(args), ensure_ascii=False)}")

    variants: list[VariantPlan] = []

    try:
        ffmpeg = resolve_binary("ffmpeg", args.ffmpeg)
        ffprobe = resolve_binary("ffprobe", args.ffprobe)
        source_paths = [Path(item).resolve() for item in args.source]
        for path in source_paths:
            if not path.is_file():
                raise SystemExit(f"Source video does not exist: {path}")
            if path.suffix.lower() not in VIDEO_SUFFIXES:
                raise SystemExit(f"Unsupported video suffix: {path}")

        rng = random.Random(args.seed or task_id)
        selected_covers: list[CoverCandidate] = []
        used_names: set[str] = set()
        total = len(source_paths) * args.target_count
        for source_index, source_path in enumerate(source_paths, start=1):
            log_event(run_log, f"source_start index={source_index} path={source_path}")
            source = load_source_video(source_path, ffprobe)
            used_signatures: set[str] = set()
            max_retries = max(args.max_retries, args.target_count * 4)
            for item_index in range(args.target_count):
                variant = build_one_variant(
                    source=source,
                    source_index=source_index,
                    item_index=item_index,
                    total=total,
                    videos_dir=videos_dir,
                    temp_cover_dir=temp_cover_dir,
                    selected_covers=selected_covers,
                    used_signatures=used_signatures,
                    used_names=used_names,
                    rng=rng,
                    max_retries=max_retries,
                    args=args,
                    ffmpeg=ffmpeg,
                )
                variants.append(variant)
                log_event(run_log, f"variant_done {len(variants)}/{total} id={variant.variant_id} output={variant.output_path}")
                print(f"[{len(variants)}/{total}] {variant.output_path}")
        payload = {
            "task_id": task_id,
            "mode": "frame_variation",
            "status": "success",
            "config": vars(args),
            "source_videos": [str(path) for path in source_paths],
            "variants": [item.to_dict() for item in variants],
        }
        write_outputs(task_root, payload, variants)
        log_event(run_log, f"success completed={len(variants)} task_root={task_root}")
        print(str(task_root))
        return 0
    except BaseException as exc:
        write_error(task_root, "frame_variation", args, variants, exc)
        raise
    finally:
        shutil.rmtree(temp_cover_dir, ignore_errors=True)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate frame-drop video variants.")
    parser.add_argument("--source", nargs="+", required=True, help="One or more source videos.")
    parser.add_argument("--output-root", required=True, help="Directory that will receive a timestamped task folder.")
    parser.add_argument("--task-name", default="campaign_batch")
    parser.add_argument("--target-count", type=int, default=20, help="Variants per source video.")
    parser.add_argument("--frames-per-second-drop", type=int, default=1)
    parser.add_argument("--width", type=int, default=720)
    parser.add_argument("--height", type=int, default=1280)
    parser.add_argument("--resize-mode", choices=("crop", "contain", "stretch", "original"), default="crop")
    parser.add_argument("--cover-hold-seconds", type=float, default=0.8)
    parser.add_argument("--max-retries", type=int, default=120)
    parser.add_argument("--max-name-length", type=int, default=240)
    parser.add_argument("--seed", default="")
    parser.add_argument("--business-tag", default="")
    parser.add_argument("--material-type", default="")
    parser.add_argument("--authorization-note", default="")
    parser.add_argument("--upload-note", default="")
    parser.add_argument("--ffmpeg", default="")
    parser.add_argument("--ffprobe", default="")
    return parser.parse_args()


def resolve_binary(name: str, explicit: str = "") -> str:
    candidates = [
        explicit,
        os.environ.get(f"{name.upper()}_BIN", ""),
        shutil.which(name) or "",
        shutil.which(f"{name}.exe") or "",
        str(Path(__file__).resolve().parent / "bin" / f"{name}.exe"),
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(Path(candidate))
    raise SystemExit(f"Missing dependency: {name}. Pass --{name} or set {name.upper()}_BIN.")


def run_process(args: list[str]) -> subprocess.CompletedProcess[str]:
    startupinfo = None
    creationflags = 0
    if hasattr(subprocess, "STARTUPINFO"):
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        creationflags = subprocess.CREATE_NO_WINDOW
    try:
        return subprocess.run(
            args,
            check=True,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            encoding="utf-8",
            errors="replace",
            startupinfo=startupinfo,
            creationflags=creationflags,
        )
    except subprocess.CalledProcessError as exc:
        raise RuntimeError((exc.stderr or exc.stdout or str(exc)).strip()) from exc


def load_source_video(path: Path, ffprobe: str) -> SourceVideo:
    result = run_process([ffprobe, "-v", "error", "-print_format", "json", "-show_streams", "-show_format", str(path)])
    data = json.loads(result.stdout)
    fmt = data.get("format", {})
    video = next((item for item in data.get("streams", []) if item.get("codec_type") == "video"), {})
    audio = next((item for item in data.get("streams", []) if item.get("codec_type") == "audio"), {})
    fps = fraction_to_float(video.get("avg_frame_rate", "0/1"))
    duration = float(fmt.get("duration", 0.0) or 0.0)
    frames = int(video.get("nb_frames", 0) or round(duration * fps))
    return SourceVideo(
        path=path,
        duration=duration,
        width=int(video.get("width", 0) or 0),
        height=int(video.get("height", 0) or 0),
        fps=fps,
        frame_count=frames,
        has_audio=bool(audio),
        bit_rate=int(fmt.get("bit_rate", 0) or 0),
        md5=compute_md5(path),
    )


def fraction_to_float(value: str) -> float:
    if "/" not in value:
        return float(value or 0.0)
    left, right = value.split("/", 1)
    denominator = float(right or 1.0)
    return float(left or 0.0) / denominator if denominator else 0.0


def compute_md5(path: Path) -> str:
    digest = hashlib.md5()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def build_one_variant(
    *,
    source: SourceVideo,
    source_index: int,
    item_index: int,
    total: int,
    videos_dir: Path,
    temp_cover_dir: Path,
    selected_covers: list[CoverCandidate],
    used_signatures: set[str],
    used_names: set[str],
    rng: random.Random,
    max_retries: int,
    args: argparse.Namespace,
    ffmpeg: str,
) -> VariantPlan:
    for attempt in range(max_retries):
        deleted_frames = choose_deleted_frames_per_second(
            source.fps,
            source.frame_count,
            args.frames_per_second_drop,
            rng,
        )
        signature = ",".join(str(frame) for frame in sorted(deleted_frames))
        if signature in used_signatures:
            continue

        output_name = unique_output_name(source.path.stem, item_index + 1, args.max_name_length, used_names)
        output_path = videos_dir / f"{output_name}.mp4"
        cover, quality, similarity = choose_cover(
            source.path,
            temp_cover_dir,
            output_name,
            selected_covers,
            source.duration,
            item_index,
            attempt,
            ffmpeg,
        )
        if similarity != "unique" and attempt < 3:
            continue
        if similarity != "unique":
            similarity = f"{similarity}_fallback"

        render_frame_drop_variant_with_cover(
            source.path,
            output_path,
            deleted_frames,
            cover.timestamp,
            args.cover_hold_seconds,
            args.width,
            args.height,
            args.resize_mode,
            ffmpeg,
        )
        selected_covers.append(cover)
        used_signatures.add(signature)
        global_index = (source_index - 1) * args.target_count + item_index + 1
        return VariantPlan(
            variant_id=f"VAR-{global_index:03d}",
            source_video=source.path,
            output_name=output_name,
            output_path=output_path,
            deleted_frames=deleted_frames,
            frame_signature=signature,
            combo_signature=signature,
            cover_timestamp=cover.timestamp,
            cover_quality_status=quality,
            cover_similarity_status=similarity,
            business_tag=args.business_tag,
            material_type=args.material_type,
            authorization_note=args.authorization_note,
            upload_note=args.upload_note,
        )
    raise RuntimeError(f"Could not find a unique frame-drop plan for {source.path.name} variant {item_index + 1}/{total}.")


def choose_deleted_frames_per_second(fps: float, frame_count: int, frames_per_second: int, rng: random.Random) -> list[int]:
    if fps <= 0 or frame_count <= 0:
        return []
    deleted: list[int] = []
    seconds = max(1, int(frame_count / fps))
    picks = max(1, frames_per_second)
    for second in range(seconds):
        start = int(round(second * fps))
        end = min(frame_count, int(round((second + 1) * fps)))
        if end <= start:
            continue
        candidates = list(range(start, end))
        deleted.extend(rng.sample(candidates, k=min(picks, len(candidates))))
    return sorted(set(deleted))


def unique_output_name(stem: str, sequence: int, max_length: int, used_names: set[str]) -> str:
    suffix = f"({sequence})"
    base = sanitize_windows_stem(stem, max(1, max_length - len(suffix)))
    candidate = f"{base}{suffix}"
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate
    offset = 2
    while True:
        alt_suffix = f"({sequence})-{offset:02d}"
        candidate = f"{sanitize_windows_stem(stem, max(1, max_length - len(alt_suffix)))}{alt_suffix}"
        if candidate not in used_names:
            used_names.add(candidate)
            return candidate
        offset += 1


def sanitize_windows_stem(name: str, max_length: int) -> str:
    invalid = set('<>:"/\\|?*')
    cleaned = "".join("_" if ch in invalid or ord(ch) < 32 else ch for ch in name.strip())
    cleaned = " ".join(cleaned.split())
    return cleaned[:max_length].rstrip(" ._-") or "variant"


def choose_cover(
    video_path: Path,
    temp_dir: Path,
    output_name: str,
    existing: list[CoverCandidate],
    duration: float,
    variant_index: int,
    attempt: int,
    ffmpeg: str,
) -> tuple[CoverCandidate, str, str]:
    usable = max(0.5, duration - 0.4)
    base = ((variant_index * 0.61803398875 + attempt * 0.17320508075) % 1.0) * usable
    timestamps = [min(max(0.2, base + offset), usable) for offset in (0.0, 0.7, 1.4)]
    candidates: list[CoverCandidate] = []
    for index, timestamp in enumerate(dict.fromkeys(round(value, 3) for value in timestamps)):
        path = temp_dir / f"{output_name}_{index}.jpg"
        extract_frame(video_path, timestamp, path, ffmpeg)
        candidates.append(evaluate_cover(path, timestamp))

    ranked = sorted(candidates, key=lambda item: (item.sharpness, -abs(128 - item.brightness)), reverse=True)
    selected = ranked[0]
    selected_quality, selected_similarity = cover_status(selected, existing)
    for candidate in ranked:
        quality, similarity = cover_status(candidate, existing)
        if quality in {"clear", "not_evaluated"} and similarity == "unique":
            selected = candidate
            selected_quality = quality
            selected_similarity = similarity
            break
    for candidate in candidates:
        if candidate.path != selected.path:
            candidate.path.unlink(missing_ok=True)
    return selected, selected_quality, selected_similarity


def extract_frame(video_path: Path, timestamp: float, output_path: Path, ffmpeg: str) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    run_process([ffmpeg, "-y", "-ss", f"{timestamp:.3f}", "-i", str(video_path), "-frames:v", "1", "-vf", "scale='min(720,iw)':-2", str(output_path)])


def evaluate_cover(path: Path, timestamp: float) -> CoverCandidate:
    try:
        from PIL import Image, ImageStat
    except ImportError:
        return CoverCandidate(timestamp=timestamp, path=path, hash_value="")

    image = Image.open(path)
    grayscale = image.convert("L")
    pixels = list(grayscale.tobytes())
    width, height = grayscale.size
    diffs: list[int] = []
    for y in range(height):
        row = y * width
        for x in range(width - 1):
            diffs.append(pixels[row + x + 1] - pixels[row + x])
    for y in range(height - 1):
        row = y * width
        next_row = (y + 1) * width
        for x in range(width):
            diffs.append(pixels[next_row + x] - pixels[row + x])
    mean = sum(diffs) / len(diffs) if diffs else 0.0
    sharpness = sum((value - mean) ** 2 for value in diffs) / len(diffs) if diffs else 0.0
    brightness = float(ImageStat.Stat(grayscale).mean[0])
    return CoverCandidate(timestamp=timestamp, path=path, sharpness=sharpness, brightness=brightness, hash_value=average_hash(image))


def average_hash(image: Any, hash_size: int = 8) -> str:
    grayscale = image.convert("L").resize((hash_size, hash_size))
    pixels = list(grayscale.tobytes())
    threshold = sum(pixels) / len(pixels)
    bits = "".join("1" if pixel > threshold else "0" for pixel in pixels)
    return f"{int(bits, 2):016x}"


def cover_status(candidate: CoverCandidate, existing: list[CoverCandidate]) -> tuple[str, str]:
    if not candidate.hash_value:
        return "not_evaluated", "unique"
    quality = "clear"
    if candidate.sharpness < 12:
        quality = "blur_risk"
    if candidate.brightness < 20 or candidate.brightness > 245:
        quality = "exposure_risk"
    similarity = "unique"
    for item in existing:
        if item.hash_value and hash_distance(candidate.hash_value, item.hash_value) < 6:
            similarity = "too_similar"
            break
    return quality, similarity


def hash_distance(left: str, right: str) -> int:
    return (int(left, 16) ^ int(right, 16)).bit_count()


def output_video_filter(width: int, height: int, mode: str) -> str:
    width = max(2, int(width or 0))
    height = max(2, int(height or 0))
    tail = "setsar=1,fps=30,format=yuv420p"
    if mode == "stretch":
        return f"scale={width}:{height},{tail}"
    if mode == "contain":
        return f"scale={width}:{height}:force_original_aspect_ratio=decrease,pad={width}:{height}:(ow-iw)/2:(oh-ih)/2:black,{tail}"
    if mode == "original":
        return tail
    return f"scale={width}:{height}:force_original_aspect_ratio=increase,crop={width}:{height},{tail}"


def render_frame_drop_variant_with_cover(
    source_path: Path,
    output_path: Path,
    deleted_frames: list[int],
    cover_timestamp: float,
    hold_seconds: float,
    width: int,
    height: int,
    resize_mode: str,
    ffmpeg: str,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    conditions = "+".join(f"eq(n\\,{frame})" for frame in sorted(set(deleted_frames)))
    drop_expr = f"not({conditions})" if conditions else "1"
    filter_complex = (
        f"[0:v:0]trim=end_frame=1,setpts=PTS-STARTPTS,tpad=stop_mode=clone:stop_duration={hold_seconds:.3f}[cover];"
        f"[1:v:0]select='{drop_expr}',setpts=N/FRAME_RATE/TB[main];"
        "[cover][main]concat=n=2:v=1:a=0[merged];"
        f"[merged]{output_video_filter(width, height, resize_mode)}[vout]"
    )
    run_process([
        ffmpeg,
        "-y",
        "-ss",
        f"{max(0.0, cover_timestamp):.3f}",
        "-i",
        str(source_path),
        "-i",
        str(source_path),
        "-filter_complex",
        filter_complex,
        "-map",
        "[vout]",
        "-map",
        "1:a?",
        "-c:v",
        "libx264",
        "-preset",
        "ultrafast",
        "-crf",
        "24",
        "-force_key_frames",
        "0",
        "-c:a",
        "copy",
        "-movflags",
        "+faststart",
        str(output_path),
    ])


def write_outputs(task_root: Path, payload: dict[str, Any], variants: list[VariantPlan]) -> None:
    task_root.mkdir(parents=True, exist_ok=True)
    (task_root / "task.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_manifest_csv(task_root / "manifest.csv", variants)
    write_manifest_xlsx(task_root / "manifest.xlsx", variants)
    write_task_log(task_root / "task.log", payload)


def write_manifest_csv(path: Path, variants: list[VariantPlan]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MANIFEST_COLUMNS)
        writer.writeheader()
        writer.writerows(plan_to_row(item) for item in variants)


def write_manifest_xlsx(path: Path, variants: list[VariantPlan]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "manifest"
    sheet.append(MANIFEST_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for item in variants:
        row = plan_to_row(item)
        sheet.append([row[column] for column in MANIFEST_COLUMNS])
    workbook.save(path)


def plan_to_row(plan: VariantPlan) -> dict[str, str]:
    return {
        "mode": plan.mode,
        "output_name": plan.output_name,
        "source_video": str(plan.source_video),
        "audio_source": "",
        "variant_id": plan.variant_id,
        "deleted_frames": ",".join(str(frame) for frame in plan.deleted_frames),
        "cover_timestamp": f"{plan.cover_timestamp:.3f}",
        "cover_quality_status": plan.cover_quality_status,
        "cover_similarity_status": plan.cover_similarity_status,
        "combo_signature": plan.combo_signature,
        "source_chain": " + ".join(str(path) for path in plan.source_chain),
        "variation_status": plan.variation_status,
        "duplicate_risk": plan.duplicate_risk,
        "quality_status": plan.quality_status,
        "business_tag": plan.business_tag,
        "material_type": plan.material_type,
        "authorization_note": plan.authorization_note,
        "upload_note": plan.upload_note,
        "creative_unit_id": plan.creative_unit_id,
        "note": plan.note,
    }


def write_task_log(path: Path, payload: dict[str, Any]) -> None:
    lines = [
        f"task_id: {payload.get('task_id', '')}",
        "mode: frame_variation",
        f"status: {payload.get('status', '')}",
        f"created_at: {datetime.now().isoformat(timespec='seconds')}",
        f"source_videos: {payload.get('source_videos', [])}",
        "",
        "[variants]",
    ]
    for item in payload.get("variants", []):
        lines.append(
            f"{item.get('variant_id')} | deleted_frames={len(item.get('deleted_frames', []))} | "
            f"cover={item.get('cover_similarity_status')}@{item.get('cover_timestamp')} | "
            f"{item.get('output_path')}"
        )
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def log_event(path: Path, message: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    timestamp = datetime.now().isoformat(timespec="seconds")
    with path.open("a", encoding="utf-8") as handle:
        handle.write(f"[{timestamp}] {message}\n")


def write_error(task_root: Path, stage: str, args: argparse.Namespace, variants: list[VariantPlan], exc: BaseException) -> None:
    payload = {
        "status": "failed",
        "stage": stage,
        "error_type": type(exc).__name__,
        "error_message": str(exc),
        "interrupted": isinstance(exc, KeyboardInterrupt),
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "config": vars(args),
        "completed_count": len(variants),
        "generated_outputs": [str(item.output_path) for item in variants],
        "variants": [item.to_dict() for item in variants],
        "traceback": traceback.format_exc(),
    }
    task_root.mkdir(parents=True, exist_ok=True)
    (task_root / "error.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    log_event(task_root / "run.log", f"failed stage={stage} type={type(exc).__name__} message={exc} completed={len(variants)}")


if __name__ == "__main__":
    raise SystemExit(main())
