from __future__ import annotations

import argparse
import csv
import itertools
import json
import os
import random
import shutil
import subprocess
from dataclasses import asdict, dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any


VIDEO_SUFFIXES = {".mp4", ".mov", ".mkv", ".avi"}
MANIFEST_COLUMNS = [
    "mode",
    "output_name",
    "source_video",
    "audio_source",
    "variant_id",
    "deleted_frames",
    "cover_timestamp",
    "cover_quality_status",
    "cover_similarity_status",
    "combo_signature",
    "source_chain",
    "variation_status",
    "duplicate_risk",
    "quality_status",
    "business_tag",
    "material_type",
    "authorization_note",
    "upload_note",
    "creative_unit_id",
    "note",
]


@dataclass
class MediaInfo:
    duration: float
    has_audio: bool


@dataclass
class VariantPlan:
    variant_id: str
    source_video: Path
    output_name: str
    output_path: Path
    source_chain: list[Path] = field(default_factory=list)
    mode: str = "folder_permutation"
    combo_signature: str = ""
    quality_status: str = "ready"
    variation_status: str = "combined"
    duplicate_risk: str = "low"
    business_tag: str = ""
    material_type: str = ""
    authorization_note: str = ""
    upload_note: str = ""
    creative_unit_id: str = ""
    note: str = ""

    def to_dict(self) -> dict[str, Any]:
        data = asdict(self)
        data["source_video"] = str(self.source_video)
        data["output_path"] = str(self.output_path)
        data["source_chain"] = [str(path) for path in self.source_chain]
        return data


def main() -> int:
    args = parse_args()
    ffmpeg = resolve_binary("ffmpeg", args.ffmpeg)
    ffprobe = resolve_binary("ffprobe", args.ffprobe)
    folders = [Path(item).resolve() for item in args.folders]
    if len(folders) < 2:
        raise SystemExit("At least two folders are required.")
    for folder in folders:
        if not folder.is_dir():
            raise SystemExit(f"Folder does not exist: {folder}")

    usage_limits = parse_usage_limits(args.folder_usage_limits, len(folders))
    rng = random.Random(args.seed or datetime.now().isoformat(timespec="microseconds"))
    chains = collect_folder_permutation_chains(folders, usage_limits, rng)
    if not chains:
        raise SystemExit("No valid folder combinations could be generated.")

    task_id = datetime.now().strftime("%Y%m%d_%H%M%S")
    task_root = Path(args.output_root).resolve() / task_id
    videos_dir = task_root / "videos"
    videos_dir.mkdir(parents=True, exist_ok=True)

    variants: list[VariantPlan] = []
    for index, chain in enumerate(chains, start=1):
        output_name = sequence_output_name(args.task_name, index, len(chains), args.max_name_length)
        output_path = videos_dir / f"{output_name}.mp4"
        render_concat_video(list(chain), output_path, args.width, args.height, args.resize_mode, ffmpeg, ffprobe)
        variant = VariantPlan(
            variant_id=f"COMBO-{index:0{max(3, len(str(len(chains))))}d}",
            source_video=chain[0],
            output_name=output_name,
            output_path=output_path,
            source_chain=list(chain),
            combo_signature=" + ".join(str(path) for path in chain),
            business_tag=args.business_tag,
            material_type=args.material_type,
            authorization_note=args.authorization_note,
            upload_note=args.upload_note,
        )
        variants.append(variant)
        print(f"[{index}/{len(chains)}] {output_path}")

    payload = {
        "task_id": task_id,
        "mode": "folder_permutation",
        "config": vars(args),
        "folders": [str(folder) for folder in folders],
        "folder_usage_limits": usage_limits,
        "variants": [item.to_dict() for item in variants],
    }
    write_outputs(task_root, payload, variants)
    print(str(task_root))
    return 0


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate ordered folder-combo videos.")
    parser.add_argument("--folders", nargs="+", required=True, help="Ordered input folders.")
    parser.add_argument("--folder-usage-limits", default="2", help="One integer or comma-separated values matching folder count.")
    parser.add_argument("--output-root", required=True)
    parser.add_argument("--task-name", default="combo_batch")
    parser.add_argument("--width", type=int, default=720)
    parser.add_argument("--height", type=int, default=1280)
    parser.add_argument("--resize-mode", choices=("crop", "contain", "stretch"), default="crop")
    parser.add_argument("--max-name-length", type=int, default=240)
    parser.add_argument("--seed", default="")
    parser.add_argument("--business-tag", default="")
    parser.add_argument("--material-type", default="")
    parser.add_argument("--authorization-note", default="")
    parser.add_argument("--upload-note", default="")
    parser.add_argument("--ffmpeg", default="")
    parser.add_argument("--ffprobe", default="")
    return parser.parse_args()


def resolve_binary(name: str, explicit: str = "") -> str:
    candidates = [
        explicit,
        os.environ.get(f"{name.upper()}_BIN", ""),
        shutil.which(name) or "",
        shutil.which(f"{name}.exe") or "",
        str(Path(__file__).resolve().parent / "bin" / f"{name}.exe"),
    ]
    for candidate in candidates:
        if candidate and Path(candidate).exists():
            return str(Path(candidate))
    raise SystemExit(f"Missing dependency: {name}. Pass --{name} or set {name.upper()}_BIN.")


def run_process(args: list[str]) -> subprocess.CompletedProcess[str]:
    startupinfo = None
    creationflags = 0
    if hasattr(subprocess, "STARTUPINFO"):
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
    if hasattr(subprocess, "CREATE_NO_WINDOW"):
        creationflags = subprocess.CREATE_NO_WINDOW
    try:
        return subprocess.run(
            args,
            check=True,
            text=True,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            encoding="utf-8",
            errors="replace",
            startupinfo=startupinfo,
            creationflags=creationflags,
        )
    except subprocess.CalledProcessError as exc:
        raise RuntimeError((exc.stderr or exc.stdout or str(exc)).strip()) from exc


def list_videos(folder: Path) -> list[Path]:
    return sorted(path for path in folder.iterdir() if path.is_file() and path.suffix.lower() in VIDEO_SUFFIXES)


def parse_usage_limits(raw: str, folder_count: int) -> list[int]:
    values = [item.strip() for item in str(raw or "").split(",") if item.strip()]
    if not values:
        return [2] * folder_count
    try:
        numbers = [max(1, int(item)) for item in values]
    except ValueError as exc:
        raise SystemExit("--folder-usage-limits must contain integers.") from exc
    if len(numbers) == 1:
        return numbers * folder_count
    if len(numbers) != folder_count:
        raise SystemExit(f"Usage limit count {len(numbers)} does not match folder count {folder_count}.")
    return numbers


def collect_folder_permutation_chains(folders: list[Path], folder_usage_limits: list[int], rng: random.Random) -> list[tuple[Path, ...]]:
    folder_videos = [list_videos(folder) for folder in folders]
    if not folder_videos or any(not videos for videos in folder_videos):
        return []
    usage_limits = {
        path: max(1, folder_usage_limits[folder_index])
        for folder_index, videos in enumerate(folder_videos)
        for path in videos
    }
    available = list(itertools.product(*folder_videos))
    target_count = combo_target_count(folder_videos, usage_limits, len(available))
    chains = build_balanced_combo_chains(folder_videos, usage_limits, target_count, rng)
    if len(chains) == target_count:
        return chains
    best_chains = chains
    for _ in range(160):
        chains = build_greedy_combo_chains(available, usage_limits, target_count, rng)
        if len(chains) == target_count:
            return chains
        if len(chains) > len(best_chains):
            best_chains = chains
    return best_chains


def combo_target_count(folder_videos: list[list[Path]], usage_limits: dict[Path, int], available_count: int) -> int:
    capacities = [sum(usage_limits[path] for path in videos) for videos in folder_videos]
    return min(available_count, min(capacities, default=0))


def build_balanced_combo_chains(
    folder_videos: list[list[Path]],
    usage_limits: dict[Path, int],
    target_count: int,
    rng: random.Random,
) -> list[tuple[Path, ...]]:
    if target_count <= 0:
        return []
    for _ in range(500):
        pools: list[list[Path]] = []
        for videos in folder_videos:
            pool = list(itertools.chain.from_iterable([path] * usage_limits[path] for path in videos))
            rng.shuffle(pool)
            pools.append(pool[:target_count])
        if any(len(pool) < target_count for pool in pools):
            return []
        chains = [tuple(pool[index] for pool in pools) for index in range(target_count)]
        if len(set(chains)) == len(chains):
            rng.shuffle(chains)
            return chains
    return []


def build_greedy_combo_chains(
    available: list[tuple[Path, ...]],
    usage_limits: dict[Path, int],
    target_count: int,
    rng: random.Random,
) -> list[tuple[Path, ...]]:
    usage_counts = {path: 0 for path in usage_limits}
    adjacent_counts: dict[tuple[int, Path, Path], int] = {}
    remaining = list(available)
    chains: list[tuple[Path, ...]] = []
    while len(chains) < target_count:
        valid = [candidate for candidate in remaining if all(usage_counts[path] < usage_limits[path] for path in candidate)]
        if not valid:
            break
        scored = [
            (
                sum(usage_counts[path] for path in candidate),
                sum(adjacent_counts.get((idx, candidate[idx], candidate[idx + 1]), 0) for idx in range(len(candidate) - 1)),
                tuple(path.name for path in candidate),
                candidate,
            )
            for candidate in valid
        ]
        scored.sort(key=lambda item: (item[0], item[1], item[2]))
        selected = rng.choice([item[3] for item in scored[: min(4, len(scored))]])
        chains.append(selected)
        for path in selected:
            usage_counts[path] += 1
        for idx in range(len(selected) - 1):
            key = (idx, selected[idx], selected[idx + 1])
            adjacent_counts[key] = adjacent_counts.get(key, 0) + 1
        remaining.remove(selected)
    return chains


def sequence_output_name(task_name: str, sequence: int, total_count: int, max_length: int) -> str:
    width = max(2, len(str(total_count)))
    suffix = f"-{sequence:0{width}d}"
    base = sanitize_windows_stem(task_name, max(1, max_length - len(suffix)))
    return f"{base}{suffix}"


def sanitize_windows_stem(name: str, max_length: int) -> str:
    invalid = set('<>:"/\\|?*')
    cleaned = "".join("_" if ch in invalid or ord(ch) < 32 else ch for ch in name.strip())
    cleaned = " ".join(cleaned.split())
    return cleaned[:max_length].rstrip(" ._-") or "variant"


def probe_media(path: Path, ffprobe: str) -> MediaInfo:
    result = run_process([ffprobe, "-v", "error", "-print_format", "json", "-show_streams", "-show_format", str(path)])
    data = json.loads(result.stdout)
    duration = float(data.get("format", {}).get("duration", 0.0) or 0.0)
    has_audio = any(item.get("codec_type") == "audio" for item in data.get("streams", []))
    return MediaInfo(duration=max(duration, 0.1), has_audio=has_audio)


def output_video_filter(width: int, height: int, mode: str) -> str:
    width = max(2, int(width or 0))
    height = max(2, int(height or 0))
    tail = "setsar=1,fps=30,format=yuv420p"
    if mode == "stretch":
        return f"scale={width}:{height},{tail}"
    if mode == "contain":
        return f"scale={width}:{height}:force_original_aspect_ratio=decrease,pad={width}:{height}:(ow-iw)/2:(oh-ih)/2:black,{tail}"
    return f"scale={width}:{height}:force_original_aspect_ratio=increase,crop={width}:{height},{tail}"


def render_concat_video(
    source_paths: list[Path],
    output_path: Path,
    width: int,
    height: int,
    resize_mode: str,
    ffmpeg: str,
    ffprobe: str,
) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    inputs: list[str] = []
    video_indices: list[int] = []
    audio_indices: list[int] = []
    input_index = 0
    for path in source_paths:
        info = probe_media(path, ffprobe)
        inputs.extend(["-i", str(path)])
        video_indices.append(input_index)
        if info.has_audio:
            audio_indices.append(input_index)
            input_index += 1
        else:
            input_index += 1
            inputs.extend(["-f", "lavfi", "-t", f"{info.duration:.3f}", "-i", "anullsrc=r=48000:cl=stereo"])
            audio_indices.append(input_index)
            input_index += 1

    parts: list[str] = []
    concat_inputs: list[str] = []
    for idx, (video_input, audio_input) in enumerate(zip(video_indices, audio_indices)):
        parts.append(f"[{video_input}:v:0]{output_video_filter(width, height, resize_mode)}[v{idx}]")
        parts.append(f"[{audio_input}:a:0]aresample=48000,aformat=sample_fmts=fltp:sample_rates=48000:channel_layouts=stereo[a{idx}]")
        concat_inputs.append(f"[v{idx}][a{idx}]")
    filter_complex = ";".join(parts) + ";" + "".join(concat_inputs) + f"concat=n={len(source_paths)}:v=1:a=1[v][a]"
    run_process([
        ffmpeg,
        "-y",
        *inputs,
        "-filter_complex",
        filter_complex,
        "-map",
        "[v]",
        "-map",
        "[a]",
        "-c:v",
        "libx264",
        "-preset",
        "fast",
        "-crf",
        "22",
        "-c:a",
        "aac",
        "-movflags",
        "+faststart",
        str(output_path),
    ])


def write_outputs(task_root: Path, payload: dict[str, Any], variants: list[VariantPlan]) -> None:
    task_root.mkdir(parents=True, exist_ok=True)
    (task_root / "task.json").write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    write_manifest_csv(task_root / "manifest.csv", variants)
    write_manifest_xlsx(task_root / "manifest.xlsx", variants)
    write_task_log(task_root / "task.log", payload)


def write_manifest_csv(path: Path, variants: list[VariantPlan]) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=MANIFEST_COLUMNS)
        writer.writeheader()
        writer.writerows(plan_to_row(item) for item in variants)


def write_manifest_xlsx(path: Path, variants: list[VariantPlan]) -> None:
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "manifest"
    sheet.append(MANIFEST_COLUMNS)
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
    for item in variants:
        row = plan_to_row(item)
        sheet.append([row[column] for column in MANIFEST_COLUMNS])
    workbook.save(path)


def plan_to_row(plan: VariantPlan) -> dict[str, str]:
    return {
        "mode": plan.mode,
        "output_name": plan.output_name,
        "source_video": str(plan.source_video),
        "audio_source": "",
        "variant_id": plan.variant_id,
        "deleted_frames": "",
        "cover_timestamp": "",
        "cover_quality_status": "n/a",
        "cover_similarity_status": "n/a",
        "combo_signature": plan.combo_signature,
        "source_chain": " + ".join(str(path) for path in plan.source_chain),
        "variation_status": plan.variation_status,
        "duplicate_risk": plan.duplicate_risk,
        "quality_status": plan.quality_status,
        "business_tag": plan.business_tag,
        "material_type": plan.material_type,
        "authorization_note": plan.authorization_note,
        "upload_note": plan.upload_note,
        "creative_unit_id": plan.creative_unit_id,
        "note": plan.note,
    }


def write_task_log(path: Path, payload: dict[str, Any]) -> None:
    lines = [
        f"task_id: {payload.get('task_id', '')}",
        "mode: folder_permutation",
        f"created_at: {datetime.now().isoformat(timespec='seconds')}",
        f"folders: {payload.get('folders', [])}",
        f"folder_usage_limits: {payload.get('folder_usage_limits', [])}",
        "",
        "[variants]",
    ]
    for item in payload.get("variants", []):
        lines.append(f"{item.get('variant_id')} | chain={item.get('combo_signature')} | {item.get('output_path')}")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    raise SystemExit(main())
