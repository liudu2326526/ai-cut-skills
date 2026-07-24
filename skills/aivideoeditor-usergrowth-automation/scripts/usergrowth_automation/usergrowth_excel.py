from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from dataclasses import dataclass
from pathlib import Path
import re
from typing import Iterable
from urllib.parse import parse_qs, urlparse
from zipfile import ZIP_DEFLATED, ZipFile
from xml.etree import ElementTree as ET

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from .usergrowth_models import UserGrowthOrderRow, UserGrowthSongRecord, UserGrowthVideoItem
from .usergrowth_rules import normalize_song_id, normalize_text

ORDER_ID_ALIASES = ("订单id", "订单ID", "订单 Id", "order_id", "orderId", "订单号")
MATERIAL_TYPE_ALIASES = ("素材类型", "类型", "功能卖点", "分类标签")
SONG_NAME_ALIASES = ("歌名", "歌曲名", "曲名", "歌曲名称")
CID_ALIASES = ("CID", "cid", "对象ID", "对象id", "creative_unit_id")
SONG_ID_ALIASES = ("标签ID", "歌曲ID", "ID", "id", "song_id", "gq", "gd")
BACKFILL_SONG_ID_ALIASES = ("标签ID", "歌曲ID", "歌曲 ID", "song_id", "gq", "gd")
SONG_LINK_ALIASES = ("链接", "歌名&链接", "歌曲链接", "song_link", "url")
ARTIST_ALIASES = ("歌手", "歌手名", "艺人", "艺人名", "演唱", "演唱者", "artist", "singer", "author")
BLOCKED_ALIASES = ("禁投", "是否禁投", "备注", "状态", "是否制作")
MISSING_SONG_ID_BACKFILL_NOTE = "未填写歌曲id自定义标签"


@dataclass
class WorkbookContext:
    """保留工作簿路径和工作簿对象的轻量上下文。"""

    path: Path
    workbook: object


def load_order_rows(path: Path) -> list[UserGrowthOrderRow]:
    """读取回填 Excel 中已有的订单、素材类型、歌曲和 CID 信息。"""
    workbook = _load_workbook(path, "回填 Excel")
    rows: list[UserGrowthOrderRow] = []
    for sheet in workbook.worksheets:
        headers = _headers(sheet)
        if not headers:
            continue
        order_col = _find_col(headers, ORDER_ID_ALIASES)
        cid_col = _find_col(headers, CID_ALIASES)
        material_col = _find_col(headers, MATERIAL_TYPE_ALIASES)
        song_col = _find_col(headers, SONG_NAME_ALIASES)
        if not order_col and not cid_col and not material_col:
            continue
        for row_number in range(2, sheet.max_row + 1):
            order_id = _cell_text(sheet, row_number, order_col)
            material_type = _cell_text(sheet, row_number, material_col)
            song_name = _cell_text(sheet, row_number, song_col)
            cid = _cell_text(sheet, row_number, cid_col)
            if not any((order_id, material_type, song_name, cid)):
                continue
            rows.append(
                UserGrowthOrderRow(
                    order_id=order_id,
                    material_type=material_type,
                    song_name=song_name,
                    cid=cid,
                    sheet_name=sheet.title,
                    row_number=row_number,
                )
            )
    return rows


def load_song_records(
        path: Path,
        duplicate_output_path: Path | None = None,
        duplicate_song_names: Iterable[str] | None = None,
) -> list[UserGrowthSongRecord]:
    """读取歌曲库 Excel，提取歌名、歌曲 ID，并按需导出当前批次涉及的同名歌曲。"""
    workbook = _load_workbook(path, "歌曲库 Excel", data_only=True)
    records: list[UserGrowthSongRecord] = []
    resolved_song_ids: dict[str, str] = {}
    for sheet in workbook.worksheets:
        header_row, headers = _song_sheet_header(sheet)
        if not header_row or not headers:
            print(f"[load_song_records] 跳过 sheet={sheet.title}，原因：未识别到歌曲名列和ID列")
            continue
        song_col = _find_col(headers, SONG_NAME_ALIASES + SONG_LINK_ALIASES)
        link_col = _find_col(headers, SONG_LINK_ALIASES)
        id_col = _find_song_id_col(headers)
        artist_col = _find_col(headers, ARTIST_ALIASES)
        blocked_cols = _find_cols(headers, BLOCKED_ALIASES)
        if not song_col:
            print(f"[load_song_records] 跳过 sheet={sheet.title}，原因：未识别到歌曲名列")
            continue
        before_count = len(records)
        skipped_missing_song = 0
        skipped_missing_id = 0
        skipped_missing_both = 0
        resolved_from_link = 0
        for row_number in range(header_row + 1, sheet.max_row + 1):
            raw_song = _cell_text(sheet, row_number, song_col)
            raw_song_id = _cell_text(sheet, row_number, id_col)
            raw_link = _cell_text(sheet, row_number, link_col)
            raw_artist = _cell_text(sheet, row_number, artist_col)
            song_link = _song_link_from_row(sheet, row_number, link_col, song_col)
            song_id = normalize_song_id(raw_song_id)
            if not song_id:
                song_id = _resolve_song_id_from_text(song_link or raw_link or raw_song, resolved_song_ids)
                if song_id:
                    resolved_from_link += 1
            if not raw_song and not raw_song_id and not raw_link:
                skipped_missing_both += 1
                continue
            if not raw_song:
                skipped_missing_song += 1
                continue
            if not song_id:
                skipped_missing_id += 1
                continue
            song_name = _extract_song_name_from_cell(raw_song)
            blocked = any("禁投" in _cell_text(sheet, row_number, col) for col in blocked_cols)
            records.append(
                UserGrowthSongRecord(
                    song_name=song_name,
                    song_id=song_id,
                    artist_name=raw_artist,
                    link=song_link,
                    blocked=blocked,
                    sheet_name=sheet.title,
                    row_number=row_number,
                )
            )
        sheet_count = len(records) - before_count
        print(
            f"[load_song_records] 读取 sheet={sheet.title}，"
            f"header_row={header_row}，新增歌曲 {sheet_count} 条，"
            f"链接补ID={resolved_from_link} 条，"
            f"跳过 缺歌名={skipped_missing_song} 条，"
            f"缺ID={skipped_missing_id} 条，"
            f"歌名、ID、链接都为空={skipped_missing_both} 条"
        )
    records, duplicate_records = _split_duplicate_song_records(records)
    duplicate_name_count = len({record.song_name for record in duplicate_records})
    if duplicate_records:
        print(
            f"[load_song_records] 跳过完全同名歌曲 {len(duplicate_records)} 条，"
            f"涉及 {duplicate_name_count} 个歌名"
        )
        if duplicate_output_path:
            export_records = _filter_duplicate_song_records_for_export(duplicate_records, duplicate_song_names)
            if export_records:
                write_duplicate_song_records(duplicate_output_path, export_records)
                print(f"[load_song_records] 同名歌曲已导出：{duplicate_output_path}")
    print(f"[load_song_records] 全部 sheet 读取完成，共 {len(records)} 条可用歌曲记录")
    return records


def _split_duplicate_song_records(
        records: list[UserGrowthSongRecord],
) -> tuple[list[UserGrowthSongRecord], list[UserGrowthSongRecord]]:
    """把完全同名歌曲从可用记录中剔除，并返回被剔除的重复记录。"""
    grouped: dict[str, list[UserGrowthSongRecord]] = defaultdict(list)
    for record in records:
        grouped[record.song_name].append(record)

    duplicate_ids = {
        id(record)
        for group in grouped.values()
        if len(group) > 1
        for record in group
    }
    unique_records = [record for record in records if id(record) not in duplicate_ids]
    duplicate_records = [record for record in records if id(record) in duplicate_ids]
    return unique_records, duplicate_records


def _filter_duplicate_song_records_for_export(
        records: list[UserGrowthSongRecord],
        song_names: Iterable[str] | None,
) -> list[UserGrowthSongRecord]:
    """只导出当前任务批次实际涉及的同名歌曲；未传过滤条件时保持全量导出。"""
    if song_names is None:
        return records
    wanted = {normalize_text(name) for name in song_names if normalize_text(name)}
    if not wanted:
        return []
    return [record for record in records if normalize_text(record.song_name) in wanted]


def write_duplicate_song_records(output_path: Path, records: list[UserGrowthSongRecord]) -> Path:
    """把被跳过的完全同名歌曲写入独立 Excel，便于人工检查歌曲库。"""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "完全同名歌曲"
    grouped: dict[str, list[UserGrowthSongRecord]] = defaultdict(list)
    for record in records:
        grouped[record.song_name].append(record)
    max_duplicate_count = max((len(group) for group in grouped.values()), default=0)

    headers = ["歌名", "重复数量"]
    for index in range(1, max_duplicate_count + 1):
        headers.extend(
            (
                f"歌手{index}",
                f"歌曲ID{index}",
                f"链接{index}",
                f"禁投{index}",
                f"来源Sheet{index}",
                f"来源行号{index}",
            )
        )
    for column, header in enumerate(headers, start=1):
        sheet.cell(row=1, column=column, value=header)

    for row_number, song_name in enumerate(sorted(grouped), start=2):
        group = sorted(grouped[song_name], key=lambda record: (record.sheet_name, record.row_number, record.song_id))
        values: list[object] = [song_name, len(group)]
        link_columns: list[int] = []
        for index, record in enumerate(group, start=1):
            values.extend(
                (
                    record.artist_name,
                    record.song_id,
                    record.link,
                    "是" if record.blocked else "",
                    record.sheet_name,
                    record.row_number,
                )
            )
            link_columns.append(2 + (index - 1) * 6 + 3)
        for column, value in enumerate(values, start=1):
            sheet.cell(row=row_number, column=column, value=value)
            if column in link_columns and value:
                sheet.cell(row=row_number, column=column).hyperlink = value

    for column in range(1, len(headers) + 1):
        if column == 1:
            width = 28
        elif column == 2:
            width = 10
        else:
            field_offset = (column - 3) % 6
            width = (18, 24, 48, 8, 18, 10)[field_offset]
        sheet.column_dimensions[sheet.cell(row=1, column=column).column_letter].width = width
    sheet.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def match_song_record(
        song_name: str,
        records: Iterable[UserGrowthSongRecord]
) -> tuple[UserGrowthSongRecord | None, list[UserGrowthSongRecord]]:
    """按歌曲名精确匹配歌曲记录，并返回候选项。"""

    records = list(records)

    # 从素材名称中提取真实歌曲名
    extracted_song_name = _extract_song_name_from_material(song_name)

    normalized = normalize_text(extracted_song_name)

    if not normalized:
        return None, []

    exact_matches = [
        item
        for item in records
        if normalize_text(item.song_name) == normalized
    ]

    # 唯一匹配直接返回
    if len(exact_matches) == 1:
        return exact_matches[0], exact_matches

    # 多个同名歌曲，交给人工确认
    if len(exact_matches) > 1:
        print("发现多个同名歌曲:")
        for item in exact_matches:
            print(
                f"歌名: {item.song_name}, "
                f"歌曲ID: {item.song_id}, "
                f"sheet: {item.sheet_name}, "
                f"row: {item.row_number}, "
                f"禁投: {item.blocked}"
            )

    return None, exact_matches


def _extract_song_name_from_material(value: str) -> str:
    """从素材名称中提取纯歌曲名。

    例如：
    dxzc-0206-507381-100427-yqf-汽水音乐-LUNA_单曲-爱我还是他
    =>
    爱我还是他
    """
    text = str(value or "").strip()

    if not text:
        return ""

    # 处理素材命名规则：LUNA_单曲-歌曲名
    if "LUNA_单曲-" in text:
        text = text.split("LUNA_单曲-", 1)[1]

    # 去掉末尾编号，例如：
    # 爱我还是他-1
    if text.endswith("-1"):
        text = text[:-2]

    return text.strip(" -_")


def write_back_results(
        order_excel: Path,
        output_path: Path,
        items: list[UserGrowthVideoItem],
        *,
        include_ready: bool = True,
) -> Path:
    """把上传成功或预检通过的素材结果写入回填 Excel。"""
    workbook = _load_workbook(order_excel, "回填 Excel")
    sheet = _select_backfill_sheet(workbook)
    backfill_items = _items_for_backfill(items, include_ready=include_ready)
    headers = _prepare_backfill_headers(sheet)
    _ensure_song_name_after_cid(sheet, headers)
    headers = _headers(sheet)
    if any(_needs_missing_song_id_note(item) for item in backfill_items):
        _ensure_header(sheet, headers, "备注", ("说明", "备注", "异常原因"))
        headers = _headers(sheet)
    columns = _backfill_columns(headers)
    if not columns:
        raise RuntimeError("回填 Excel 中没有可写入的列，请保留素材类型、时间、CID、类型等表头。")

    row_number = _first_empty_backfill_row(sheet, columns)
    for item in backfill_items:
        while _row_has_existing_cid(sheet, row_number, columns):
            row_number += 1
        _write_backfill_row(sheet, row_number, columns, item)
        row_number += 1

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.active = workbook.worksheets.index(sheet)
    workbook.save(output_path)
    return output_path


def _items_for_backfill(items: list[UserGrowthVideoItem], *, include_ready: bool = True) -> list[UserGrowthVideoItem]:
    """筛出需要写入回填表的素材条目。"""
    statuses = {"success", "ready"} if include_ready else {"success"}
    return [item for item in items if item.status in statuses]


def _needs_missing_song_id_note(item: UserGrowthVideoItem) -> bool:
    """判断回填行是否需要提示歌曲 ID 未写入自定义标签。"""
    if item.material_type in {"金币VIP", "金币SVIP"}:
        return False
    return not str(item.song_id or "").strip()


def _select_backfill_sheet(workbook) -> Worksheet:
    """从工作簿中选择最像回填模板的 sheet。"""
    for sheet in workbook.worksheets:
        headers = _headers(sheet)
        if _find_col(headers, CID_ALIASES) or _find_col(headers, MATERIAL_TYPE_ALIASES):
            return sheet
    return workbook.worksheets[0]


def _prepare_backfill_headers(sheet: Worksheet) -> dict[str, int]:
    """读取表头；如果模板完全空白，则创建最小回填表头。"""
    headers = _headers(sheet)
    if headers:
        return headers

    for column, header in enumerate(("素材类型", "时间", "CID", "类型"), start=1):
        sheet.cell(row=1, column=column, value=header)
    return _headers(sheet)


def _ensure_song_name_after_cid(sheet: Worksheet, headers: dict[str, int]) -> int | None:
    """确保 CID 后面有歌曲名称列，用于回填当前文件对应的歌曲名。"""
    existing_song_col = _find_col(headers, SONG_NAME_ALIASES)
    if existing_song_col:
        return existing_song_col

    cid_col = _find_col(headers, CID_ALIASES)
    if not cid_col:
        return None

    insert_col = cid_col + 1
    sheet.insert_cols(insert_col)
    sheet.cell(row=1, column=insert_col, value="歌曲名称")
    return insert_col


def _backfill_columns(headers: dict[str, int]) -> dict[str, int]:
    """把不同名称的表头映射到统一的回填字段名。"""
    mapping = {
        "cid": _find_col(headers, CID_ALIASES),
        "material_type": _find_col(headers, MATERIAL_TYPE_ALIASES),
        "type": _find_col(headers, ("类型",)),
        "time": _find_col(headers, ("时间",)),
        "order_id": _find_col(headers, ORDER_ID_ALIASES),
        "song_name": _find_col(headers, SONG_NAME_ALIASES),
        "song_id": _find_col(headers, BACKFILL_SONG_ID_ALIASES),
        "file_name": _find_col(headers, ("文件名", "素材名称", "视频名称")),
        "status": _find_col(headers, ("状态",)),
        "message": _find_col(headers, ("说明", "备注", "异常原因")),
        "classification_path": _find_col(headers, ("分类标签",)),
        "custom_tags": _find_col(headers, ("自定义标签",)),
        "optional_tags": _find_col(headers, ("选填标签",)),
    }
    columns: dict[str, int] = {}
    used: set[int] = set()
    for key, column in mapping.items():
        if column and column not in used:
            columns[key] = column
            used.add(column)
    return columns


def _first_empty_backfill_row(sheet: Worksheet, columns: dict[str, int]) -> int:
    """找到下一行可写入回填结果的位置。"""
    cid_col = columns.get("cid")
    if cid_col:
        return _first_empty_cid_row(sheet, cid_col)

    check_columns = [columns[key] for key in ("material_type", "type", "song_name", "file_name") if key in columns]
    if not check_columns:
        return sheet.max_row + 1

    row = 2
    while row <= sheet.max_row and any(_cell_text(sheet, row, column) for column in check_columns):
        row += 1
    return row


def _row_has_existing_cid(sheet: Worksheet, row_number: int, columns: dict[str, int]) -> bool:
    """判断某一行是否已经有 CID，已有 CID 的行不能覆盖。"""
    cid_col = columns.get("cid")
    return bool(cid_col and _cell_text(sheet, row_number, cid_col))


def _write_backfill_row(sheet: Worksheet, row_number: int, columns: dict[str, int], item: UserGrowthVideoItem) -> None:
    """把单个素材的 CID、素材类型、标签和状态写入指定行。"""
    values = {
        "cid": item.cid,
        "material_type": item.cid_material_type or item.material_type,
        "type": "剪辑",
        "time": "",
        "order_id": item.order_id,
        "song_name": item.song_name,
        "song_id": item.song_id,
        "file_name": item.file_name,
        "status": item.status,
        "message": _backfill_message(item),
        "classification_path": " / ".join(item.classification_path),
        "custom_tags": "、".join(item.custom_tags),
        "optional_tags": "、".join(item.optional_tags),
    }
    for key, value in values.items():
        column = columns.get(key)
        if column and (value or key in {"time", "cid"}):
            sheet.cell(row=row_number, column=column, value=value)


def _backfill_message(item: UserGrowthVideoItem) -> str:
    """生成回填备注；歌曲 ID 为空时追加自定义标签缺失提示。"""
    message = str(item.message or "").strip()
    if not _needs_missing_song_id_note(item):
        return message
    if MISSING_SONG_ID_BACKFILL_NOTE in message:
        return message
    return f"{message}；{MISSING_SONG_ID_BACKFILL_NOTE}" if message else MISSING_SONG_ID_BACKFILL_NOTE


def _load_workbook(path: Path, label: str, **kwargs):
    """读取 Excel；遇到样式损坏时尝试自动剥离坏样式后再读取。"""
    try:
        return load_workbook(path, **kwargs)
    except Exception as exc:  # noqa: BLE001
        try:
            return _load_workbook_with_repaired_styles(path, **kwargs)
        except Exception as repair_exc:  # noqa: BLE001
            repair_note = f"\n自动修复样式也失败：{repair_exc}"
        raise RuntimeError(
            f"{label} 读取失败：{path}\n"
            "建议用 Excel 或 WPS 打开这个文件，另存为 .xlsx 后再试。\n"
            f"原始错误：{exc}{repair_note}"
        ) from exc


def _load_workbook_with_repaired_styles(path: Path, **kwargs):
    """使用修复后的内存字节流重新加载工作簿。"""
    repaired_data = _workbook_bytes_with_repaired_styles(path)
    return load_workbook(BytesIO(repaired_data), **kwargs)


def _workbook_bytes_with_repaired_styles(path: Path) -> bytes:
    """重写 xlsx/xlsm 中的样式文件，绕开 openpyxl 无法解析的坏样式。"""
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        raise ValueError("仅支持自动修复 .xlsx/.xlsm 文件")

    output = BytesIO()
    with ZipFile(path, "r") as source, ZipFile(output, "w", ZIP_DEFLATED) as target:
        seen: set[str] = set()
        found_styles = False
        for info in source.infolist():
            if info.filename in seen:
                continue
            seen.add(info.filename)
            with source.open(info) as item:
                data = item.read()
            if info.filename == "xl/styles.xml":
                data = _minimal_styles_xml()
                found_styles = True
            elif info.filename.startswith("xl/worksheets/") and info.filename.endswith(".xml"):
                data = _strip_worksheet_style_dependencies(data)
            target.writestr(info, data)
        if not found_styles:
            raise ValueError("文件里没有找到样式定义")
    return output.getvalue()


def _minimal_styles_xml() -> bytes:
    """生成 openpyxl 可接受的最小 styles.xml。"""
    return b"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<styleSheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
  <fonts count="1">
    <font><sz val="11"/><color theme="1"/><name val="Calibri"/><family val="2"/><scheme val="minor"/></font>
  </fonts>
  <fills count="2">
    <fill><patternFill patternType="none"/></fill>
    <fill><patternFill patternType="gray125"/></fill>
  </fills>
  <borders count="1"><border><left/><right/><top/><bottom/><diagonal/></border></borders>
  <cellStyleXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0"/></cellStyleXfs>
  <cellXfs count="1"><xf numFmtId="0" fontId="0" fillId="0" borderId="0" xfId="0"/></cellXfs>
  <cellStyles count="1"><cellStyle name="Normal" xfId="0" builtinId="0"/></cellStyles>
  <dxfs count="0"/>
  <tableStyles count="0" defaultTableStyle="TableStyleMedium9" defaultPivotStyle="PivotStyleLight16"/>
</styleSheet>"""


def _strip_worksheet_style_dependencies(data: bytes) -> bytes:
    """移除 worksheet XML 中对坏样式和条件格式的引用。"""
    root = ET.fromstring(data)
    namespace = root.tag[1:].split("}", 1)[0] if root.tag.startswith("{") else ""
    conditional_formatting_tag = f"{{{namespace}}}conditionalFormatting" if namespace else "conditionalFormatting"
    changed = False

    for element in root.iter():
        for attribute in ("s", "style", "customFormat"):
            if attribute in element.attrib:
                element.attrib.pop(attribute, None)
                changed = True

    for child in list(root):
        if child.tag == conditional_formatting_tag:
            root.remove(child)
            changed = True

    if not changed:
        return data
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def _headers(sheet: Worksheet) -> dict[str, int]:
    """读取第一行表头，返回表头文本到列号的映射。"""
    values: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        value = _cell_text(sheet, 1, column)
        if value:
            values[value] = column
    return values


def _find_col(headers: dict[str, int], aliases: Iterable[str]) -> int | None:
    """按别名查找最匹配的列号。"""
    normalized = {normalize_text(name): column for name, column in headers.items()}
    for alias in aliases:
        column = normalized.get(normalize_text(alias))
        if column:
            return column
    for name, column in headers.items():
        normalized_name = normalize_text(name)
        if any(normalize_text(alias) in normalized_name for alias in aliases):
            return column
    return None


def _find_song_id_col(headers: dict[str, int]) -> int | None:
    """优先按既有别名找歌曲 ID 列，兜底匹配任意包含 id 的表头。"""
    column = _find_col(headers, SONG_ID_ALIASES)
    if column:
        return column
    for name, candidate in headers.items():
        if "id" in str(name or "").lower():
            return candidate
    return None


def _song_sheet_header(sheet: Worksheet) -> tuple[int, dict[str, int]]:
    """在歌曲库 sheet 的前几行里定位真正的表头行。"""
    max_header_row = min(sheet.max_row, 50)
    for row_number in range(1, max_header_row + 1):
        headers = _headers_at_row(sheet, row_number)
        if not headers:
            continue
        song_col = _find_col(headers, SONG_NAME_ALIASES + SONG_LINK_ALIASES)
        id_col = _find_song_id_col(headers)
        link_col = _find_col(headers, SONG_LINK_ALIASES)
        if song_col and (id_col or link_col):
            return row_number, headers
    return 0, {}


def _headers_at_row(sheet: Worksheet, row_number: int) -> dict[str, int]:
    """读取指定行表头，返回表头文本到列号的映射。"""
    values: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        value = _cell_text(sheet, row_number, column)
        if value:
            values[value] = column
    return values


def _find_cols(headers: dict[str, int], aliases: Iterable[str]) -> list[int]:
    """按多个别名查找所有可能的列号。"""
    columns: list[int] = []
    for alias in aliases:
        col = _find_col(headers, (alias,))
        if col and col not in columns:
            columns.append(col)
    return columns


def _ensure_header(sheet: Worksheet, headers: dict[str, int], preferred: str, aliases: Iterable[str]) -> int:
    """确保某个字段存在表头，不存在时追加新列。"""
    column = _find_col(headers, aliases)
    if column:
        return column
    column = 1 if sheet.max_column == 1 and not _cell_text(sheet, 1, 1) else sheet.max_column + 1
    sheet.cell(row=1, column=column, value=preferred)
    return column


def _first_empty_cid_row(sheet: Worksheet, cid_col: int) -> int:
    """从第二行开始找到 CID 列为空的第一行。"""
    row = 2
    while row <= sheet.max_row and _cell_text(sheet, row, cid_col):
        row += 1
    return row


def _cell_text(sheet: Worksheet, row: int, col: int | None) -> str:
    """读取单元格文本，空列或空值统一返回空字符串。"""
    if not col:
        return ""
    value = sheet.cell(row=row, column=col).value
    return str(value or "").strip()


def _cell_hyperlink(sheet: Worksheet, row: int, col: int | None) -> str:
    """读取单元格超链接地址，空列或无超链接时返回空字符串。"""
    if not col:
        return ""
    hyperlink = sheet.cell(row=row, column=col).hyperlink
    target = getattr(hyperlink, "target", "") if hyperlink else ""
    return str(target or "").strip()


def _song_link_from_row(sheet: Worksheet, row: int, link_col: int | None, song_col: int | None) -> str:
    """优先读取链接列超链接，再从链接列/歌名单元格文本中提取 URL。"""
    for column in (link_col, song_col):
        link = _cell_hyperlink(sheet, row, column)
        if link:
            return link
    for text in (_cell_text(sheet, row, link_col), _cell_text(sheet, row, song_col)):
        link = _extract_first_url(text)
        if link:
            return link
    return ""


def _extract_song_name_from_cell(value: str) -> str:
    """从歌曲库的“歌名&链接”等混合文本中提取纯歌名。"""
    text = str(value or "").strip()
    if "《" in text and "》" in text:
        start = text.find("《") + 1
        end = text.find("》", start)
        if end > start:
            return text[start:end].strip()
    if "@" in text:
        text = text.split("@", 1)[0].strip()
    return text.strip("《》 ")


def _resolve_song_id_from_text(text: str, cache: dict[str, str] | None = None) -> str:
    """从文本中的分享链接解析歌曲 ID。"""
    link = _extract_first_url(text)
    if not link:
        return ""
    if cache is not None and link in cache:
        return cache[link]
    track_id = _resolve_track_id_from_share_link(link)
    song_id = normalize_song_id(track_id)
    if cache is not None:
        cache[link] = song_id
    return song_id


def _extract_first_url(text: str) -> str:
    """从单元格文本中提取第一个 URL。"""
    match = re.search(r"https?://[^\s]+", str(text or ""))
    return match.group(0).strip() if match else ""


def _resolve_track_id_from_share_link(url: str) -> str:
    """通过 HTTP 跟随重定向，从最终地址中提取 track_id。"""
    try:
        response = requests.get(
            url,
            allow_redirects=True,
            timeout=10,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) "
                    "Chrome/137.0.0.0 Safari/537.36"
                )
            },
        )
    except Exception:
        return ""
    final_url = response.url or ""
    query = parse_qs(urlparse(final_url).query)
    values = query.get("track_id") or query.get("trackId")
    return str(values[0]).strip() if values else ""


def _replace_sheet(workbook, title: str):
    """删除同名 sheet 后重新创建，用于需要重建结果页的场景。"""
    if title in workbook.sheetnames:
        sheet = workbook[title]
        workbook.remove(sheet)
    sheet = workbook.create_sheet(title, 0)
    workbook.active = 0
    return sheet
